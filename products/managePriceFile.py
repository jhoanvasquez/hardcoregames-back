import glob
import os

import openpyxl
from django import forms
from django.conf import settings
from django.db.models import F,Q

from products.models import ProductAccounts, Consoles, Licenses, Products, GameDetail, TypeAccounts, \
    PriceForSuscription, TypeSuscriptionAccounts


def read_file_ps(sheetPs, id_primaria, id_secundaria):
    id_ps4 = Consoles.objects.filter(descripcion__icontains="playstation 4")
    id_ps5 = Consoles.objects.filter(descripcion__icontains="playstation 5")

    sheet = sheetPs
    m_row = sheet.max_row
    batch_size = 4

    for start in range(2, m_row + 1, batch_size):
        end = min(start + batch_size, m_row + 1)
        process_batch_ps(sheet, start, end, id_ps4, id_ps5, id_primaria, id_secundaria)

def process_batch_ps(sheet, start, end, id_ps4, id_ps5, id_primaria, id_secundaria):
    for i in range(start, end):
        account = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        id_product = sheet.cell(row=i, column=3).value
        duration_days = sheet.cell(row=i, column=8).value or 0
        type_account = sheet.cell(row=i, column=9).value or 1

        if not account or not id_product:
            continue  # Skip invalid rows

        product_for_create = Products.objects.filter(id_product=id_product).first()
        if not product_for_create:
            raise forms.ValidationError(f"El producto para PlayStation con ID {id_product} no existe")

        type_account_selected = TypeAccounts.objects.filter(pk=type_account).first()

        account_for_producto, _ = ProductAccounts.objects.update_or_create(
            cuenta=account.lower(),
            defaults={
                "password": password,
                "activa": True,
                "tipo_cuenta": type_account_selected,
                "dias_duracion": duration_days,
            }
        )

        prices = {
            "ps4_1": sheet.cell(row=i, column=4).value,
            "ps4_2": sheet.cell(row=i, column=5).value,
            "ps5_1": sheet.cell(row=i, column=6).value,
            "ps5_2": sheet.cell(row=i, column=7).value,
        }

        for key, price in prices.items():
            if check_sheet_price(str(price).strip()):
                console = id_ps4 if "ps4" in key else id_ps5
                tipo = id_primaria if "1" in key else id_secundaria
                save_or_update_game_detail(id_product, console, tipo, duration_days, account_for_producto)

def read_file_xbx(sheetPs, id_primaria, id_secundaria):
    id_xbox = Consoles.objects.filter(descripcion__exact="xbox")
    id_code = Licenses.objects.filter(descripcion__icontains="codigo")
    id_pc = Consoles.objects.filter(descripcion__exact="Pc")
    licence_pc = Licenses.objects.filter(descripcion__icontains="pc")

    sheet = sheetPs
    m_row = sheet.max_row
    batch_size = 5  # Process 5 rows at a time

    for start in range(2, m_row + 1, batch_size):
        end = min(start + batch_size, m_row + 1)
        process_batch_xbx(sheet, start, end, id_xbox, id_code, id_pc, licence_pc, id_primaria, id_secundaria)

def process_batch_xbx(sheet, start, end, id_xbox, id_code, id_pc, licence_pc, id_primaria, id_secundaria):
    for i in range(start, end):
        account = sheet.cell(row=i, column=1).value
        password = sheet.cell(row=i, column=2).value
        id_product = sheet.cell(row=i, column=3).value
        duration_days = sheet.cell(row=i, column=8).value or 0

        if not account or not id_product:
            continue  # Skip invalid rows

        product_for_create = Products.objects.filter(id_product=id_product).first()
        if not product_for_create:
            raise forms.ValidationError(f"El producto para Xbox con ID {id_product} no existe")

        sheet_prices = {
            "xbox_1": str(sheet.cell(row=i, column=4).value).strip(),
            "xbox_2": str(sheet.cell(row=i, column=5).value).strip(),
            "pc": str(sheet.cell(row=i, column=6).value).strip(),
            "code": str(sheet.cell(row=i, column=7).value).strip(),
        }

        account_for_producto, _ = ProductAccounts.objects.update_or_create(
            cuenta=account.lower(),
            defaults={
                "password": password,
                "activa": True,
                "tipo_cuenta": TypeAccounts.objects.filter(pk=1).first(),
                "dias_duracion": duration_days,
            }
        )

        for key, console, license_type in [
            ("xbox_1", id_xbox, id_primaria),
            ("xbox_2", id_xbox, id_secundaria),
            ("pc", id_pc, licence_pc),
            ("code", id_xbox, id_code),
        ]:
            if check_sheet_price(sheet_prices[key]):
                save_or_update_game_detail(id_product, console, license_type, duration_days, account_for_producto)

class ManegePricesFile:
    def __init__(self):
        path_file_upload = glob.glob(settings.STATIC_URL_FILES + "/*.xlsx")[0]
        path = os.path.abspath(path_file_upload.replace('\\', '/'))
        excel_document = openpyxl.load_workbook(path)
        sheet_ps = excel_document.get_sheet_by_name('cuentas_ps')
        sheet_xbox = excel_document.get_sheet_by_name('cuentas_xbox')
        id_primaria = Licenses.objects.filter(descripcion__icontains="primaria")
        id_secundaria = Licenses.objects.filter(descripcion__icontains="secundaria")

        read_file_ps(sheet_ps, id_primaria, id_secundaria)
        read_file_xbx(sheet_xbox, id_primaria, id_secundaria)



def save_or_update_game_detail(id_product, id_console, id_license, duration_days, account):
    existing_game_detail = GameDetail.objects.filter(
                                producto_id=id_product,
                                licencia=id_license.first(),
                                consola=id_console.first(),
                                duracion_dias_alquiler=duration_days,
                            ).filter(
                                Q(precio__gt=0) | Q(precio_descuento__gt=0)
                            ).first()

    if existing_game_detail:
        price = existing_game_detail.precio
        offer_price = existing_game_detail.precio_descuento
    else:
        price = 0
        offer_price = 0

    row_game_detail, created = GameDetail.objects.get_or_create(
        producto_id=id_product,
        licencia=id_license.first(),
        consola=id_console.first(),
        duracion_dias_alquiler=duration_days,
        cuenta=account,
        defaults={
            "stock": 1,
            "precio": price,
            "precio_descuento": offer_price
        }
    )

    if not created:
        row_game_detail.stock += 1
        row_game_detail.save()


def check_sheet_price(sheet):
    return sheet.strip() != "None" or "x" in sheet